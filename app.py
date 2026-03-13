# app.py
from flask import Flask, render_template, request, redirect, url_for, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from datetime import datetime
import pandas as pd
import json
import os
import re
from werkzeug.utils import secure_filename

app = Flask(__name__)
CORS(app)

basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'shipping_cycle.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# ══════════════════════════════════════════════════════
#  MODELS
# ══════════════════════════════════════════════════════

class ShippingEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sr_no = db.Column(db.String(20), unique=True)
    entry_date = db.Column(db.String(20))
    customer_name = db.Column(db.String(100))
    container_no = db.Column(db.String(50))
    seal_no = db.Column(db.String(20))
    driver_name = db.Column(db.String(100))
    driver_phone = db.Column(db.String(20))
    license_no = db.Column(db.String(50))
    destination = db.Column(db.String(100))
    vehicle_no = db.Column(db.String(50))
    transporter = db.Column(db.String(100))
    arrival_time = db.Column(db.String(20))
    out_date_time = db.Column(db.String(50))
    status = db.Column(db.String(20), default="Open")

class Checklist7Point(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sr_no = db.Column(db.String(20), unique=True)
    checklist_json = db.Column(db.Text)
    rejection_remark = db.Column(db.Text)
    is_rejected = db.Column(db.Boolean, default=False)

class PackingList(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sr_no = db.Column(db.String(20))
    element_id = db.Column(db.String(50), unique=True)
    product_name = db.Column(db.String(255))
    qty_m2 = db.Column(db.Float)
    qty_sheets = db.Column(db.Integer)
    net_weight = db.Column(db.Float)
    gross_weight = db.Column(db.Float)
    customer_po = db.Column(db.String(50))
    invoice_no = db.Column(db.String(50))

class PalletVerification(db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    sr_no          = db.Column(db.String(20), nullable=False)
    element_id     = db.Column(db.String(50), nullable=False)
    checklist_json = db.Column(db.Text, default='[]')
    verified_at    = db.Column(db.String(30))
    __table_args__ = (
        db.UniqueConstraint('sr_no', 'element_id', name='uq_pv_sr_element'),
    )

class ShipmentPhoto(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    sr_no       = db.Column(db.String(20), nullable=False)
    file_path   = db.Column(db.String(500), nullable=False)
    category    = db.Column(db.String(50), default='Loading Photos')
    uploaded_at = db.Column(db.String(30))


# ══════════════════════════════════════════════════════
#  ROUTES — Dashboard
# ══════════════════════════════════════════════════════

@app.route('/')
def dashboard():
    entries = ShippingEntry.query.all()
    return render_template('dashboard.html', entries=entries)


# ══════════════════════════════════════════════════════
#  ROUTES — Entry
# ══════════════════════════════════════════════════════

@app.route('/api/add_entry', methods=['POST'])
def api_add_entry():
    data = request.json
    try:
        new_entry = ShippingEntry(
            sr_no=data.get('sr_no'),
            entry_date=data.get('entry_date'),
            customer_name=data.get('customer_name'),
            container_no=data.get('container_no'),
            driver_name=data.get('driver_name'),
            driver_phone=data.get('driver_phone'),
            license_no=data.get('license_no'),
            destination=data.get('destination'),
            vehicle_no=data.get('vehicle_no'),
            transporter=data.get('transporter'),
            arrival_time=data.get('arrival_time'),
            out_date_time=data.get('out_date_time'),
            status="Open"
        )
        db.session.add(new_entry)
        db.session.commit()
        return jsonify({"status": "success", "message": "Entry added"}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": "Duplicate Sr. No or Database Error"}), 400


# ══════════════════════════════════════════════════════
#  ROUTES — Step 1: 7-Point Checklist
# ══════════════════════════════════════════════════════

@app.route('/checklist/<sr_no>')
def checklist(sr_no):
    entry_data  = ShippingEntry.query.filter_by(sr_no=sr_no).first_or_404()
    saved_entry = Checklist7Point.query.filter_by(sr_no=sr_no).first()
    return render_template('7_point_checklist.html', data=entry_data, saved_data=saved_entry)

@app.route('/save_checklist', methods=['POST'])
def save_checklist():
    data        = request.json
    sr_no       = data.get('sr_no')
    is_rejected = data.get('is_rejected', False)

    check_entry = Checklist7Point.query.filter_by(sr_no=sr_no).first()
    if not check_entry:
        check_entry = Checklist7Point(sr_no=sr_no)
        db.session.add(check_entry)

    check_entry.checklist_json   = json.dumps(data.get('form_data'))
    check_entry.is_rejected      = is_rejected
    check_entry.rejection_remark = data.get('remark')

    main_entry = ShippingEntry.query.filter_by(sr_no=sr_no).first()
    if main_entry:
        updates = data.get('shipping_updates', {})
        main_entry.customer_name = updates.get('customer_name')
        main_entry.container_no  = updates.get('container_no')
        main_entry.seal_no       = updates.get('seal_no')
        main_entry.driver_name   = updates.get('driver_name')
        main_entry.driver_phone  = updates.get('driver_phone')
        main_entry.license_no    = updates.get('license_no')
        main_entry.destination   = updates.get('destination')
        main_entry.vehicle_no    = updates.get('vehicle_no')
        main_entry.transporter   = updates.get('transporter')
        main_entry.arrival_time  = updates.get('arrival_time')
        if is_rejected:
            main_entry.status = "Rejected"

    db.session.commit()
    return jsonify({"status": "success", "message": "Data saved successfully"})


# ══════════════════════════════════════════════════════
#  ROUTES — Step 2: Packing List
# ══════════════════════════════════════════════════════

@app.route('/packing_list/<sr_no>')
def packing_list_page(sr_no):
    entry_data   = ShippingEntry.query.filter_by(sr_no=sr_no).first_or_404()
    items        = PackingList.query.filter_by(sr_no=sr_no).all()
    verified_ids = [pv.element_id for pv in PalletVerification.query.filter_by(sr_no=sr_no).all()]

    total_qty_m2       = sum(i.qty_m2      or 0 for i in items)
    total_qty_sheets   = sum(i.qty_sheets  or 0 for i in items)
    total_net_weight   = sum(i.net_weight  or 0 for i in items)
    total_gross_weight = sum(i.gross_weight or 0 for i in items)

    return render_template('packing_list.html',
                           data=entry_data,
                           items=items,
                           verified_ids=verified_ids,
                           total_qty_m2=round(total_qty_m2, 3),
                           total_qty_sheets=total_qty_sheets,
                           total_net_weight=round(total_net_weight, 3),
                           total_gross_weight=round(total_gross_weight, 3))

@app.route('/upload_packing_excel/<sr_no>', methods=['POST'])
def upload_packing_excel(sr_no):
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "No file part"}), 400

    file = request.files['file']
    try:
        import openpyxl
        # Load workbook directly — faster than pandas for simple reads
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        # Read header row
        headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        required = ['Element ID', 'Product Name', 'Qty M2', 'Qty Sheets', 'Net Weight', 'Gross Weight', 'Customer PO', 'Invoice No']
        for col in required:
            if col not in headers:
                return jsonify({"status": "error", "message": f"Missing column: '{col}' in Excel file."}), 400

        idx = {col: headers.index(col) for col in required}

        # Read all rows into memory first (fast)
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            eid = row[idx['Element ID']]
            if eid is None or str(eid).strip() == '':
                continue  # skip empty rows
            rows.append(row)
        wb.close()

        if not rows:
            return jsonify({"status": "error", "message": "No data rows found in Excel."}), 400

        # Check duplicates within Excel
        eids_in_file = [str(r[idx['Element ID']]).strip() for r in rows]
        if len(eids_in_file) != len(set(eids_in_file)):
            dupes = list({x for x in eids_in_file if eids_in_file.count(x) > 1})
            return jsonify({"status": "error", "message": f"Duplicate Element IDs in Excel: {dupes}"}), 400

        # Check duplicates in DB (bulk query — much faster than one-by-one)
        existing = PackingList.query.filter(PackingList.element_id.in_(eids_in_file)).first()
        if existing:
            return jsonify({"status": "error",
                            "message": f"Element ID '{existing.element_id}' already exists in Database."}), 400

        # Bulk insert
        def safe_float(v):
            try: return float(v) if v not in (None, '') else 0.0
            except: return 0.0
        def safe_int(v):
            try: return int(float(v)) if v not in (None, '') else 0
            except: return 0

        items = [PackingList(
            sr_no=sr_no,
            element_id=str(rows[i][idx['Element ID']]).strip(),
            product_name=str(rows[i][idx['Product Name']] or ''),
            qty_m2=safe_float(rows[i][idx['Qty M2']]),
            qty_sheets=safe_int(rows[i][idx['Qty Sheets']]),
            net_weight=safe_float(rows[i][idx['Net Weight']]),
            gross_weight=safe_float(rows[i][idx['Gross Weight']]),
            customer_po=str(rows[i][idx['Customer PO']] or ''),
            invoice_no=str(rows[i][idx['Invoice No']] or '')
        ) for i in range(len(rows))]

        db.session.bulk_save_objects(items)
        db.session.commit()
        return jsonify({"status": "success", "message": f"{len(items)} item(s) imported successfully!"})

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/delete_packing_list/<sr_no>', methods=['DELETE'])
def delete_packing_list(sr_no):
    try:
        verified_count = PalletVerification.query.filter_by(sr_no=sr_no).count()
        if verified_count > 0:
            return jsonify({"status": "error",
                            "message": f"Cannot delete: {verified_count} element(s) already verified."}), 400
        deleted = PackingList.query.filter_by(sr_no=sr_no).delete()
        db.session.commit()
        return jsonify({"status": "success", "message": f"{deleted} item(s) deleted from Packing List."})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


# ══════════════════════════════════════════════════════
#  ROUTES — Step 3: Pallet Verification
# ══════════════════════════════════════════════════════

@app.route('/pallet_packing/<sr_no>')
def pallet_packing(sr_no):
    entry_data  = ShippingEntry.query.filter_by(sr_no=sr_no).first_or_404()
    all_items   = PackingList.query.filter_by(sr_no=sr_no).all()
    total_items = len(all_items)

    saved_list = []
    for pv in PalletVerification.query.filter_by(sr_no=sr_no).all():
        item = PackingList.query.filter_by(sr_no=sr_no, element_id=pv.element_id).first()
        saved_list.append({
            "element_id":   pv.element_id,
            "checklist":    json.loads(pv.checklist_json) if pv.checklist_json else [],
            "product_name": item.product_name  if item else "",
            "qty_m2":       item.qty_m2        if item else "",
            "qty_sheets":   item.qty_sheets    if item else "",
            "net_weight":   item.net_weight    if item else "",
            "gross_weight": item.gross_weight  if item else "",
            "customer_po":  item.customer_po   if item else "",
            "invoice_no":   item.invoice_no    if item else "",
        })

    return render_template('pallet_verification.html',
                           data=entry_data,
                           saved_pallets=saved_list,
                           total_items=total_items)

@app.route('/api/check_element/<sr_no>/<path:element_id>')
def check_element(sr_no, element_id):
    item = PackingList.query.filter_by(sr_no=sr_no, element_id=element_id).first()
    if not item:
        return jsonify({"status": "error",
                        "message": f"Element ID '{element_id}' is not in Packing List."}), 404
    already = PalletVerification.query.filter_by(sr_no=sr_no, element_id=element_id).first()
    return jsonify({
        "status":        "success",
        "already_saved": bool(already),
        "element_id":    item.element_id,
        "product_name":  item.product_name,
        "qty_m2":        item.qty_m2,
        "qty_sheets":    item.qty_sheets,
        "net_weight":    item.net_weight,
        "gross_weight":  item.gross_weight,
        "customer_po":   item.customer_po,
        "invoice_no":    item.invoice_no,
    })

@app.route('/api/verified_ids/<sr_no>')
def api_verified_ids(sr_no):
    ids = [pv.element_id for pv in PalletVerification.query.filter_by(sr_no=sr_no).all()]
    return jsonify({"status": "success", "verified_ids": ids})

@app.route('/save_pallet_verification/<sr_no>', methods=['POST'])
def save_pallet_verification(sr_no):
    payload = request.get_json(force=True)
    if not payload:
        return jsonify({"status": "error", "message": "No data received"}), 400
    try:
        for item in payload:
            eid       = item.get('element_id', '').strip()
            checklist = item.get('checklist', [])
            if not eid:
                continue
            pv = PalletVerification.query.filter_by(sr_no=sr_no, element_id=eid).first()
            if pv:
                pv.checklist_json = json.dumps(checklist)
                pv.verified_at    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            else:
                db.session.add(PalletVerification(
                    sr_no=sr_no, element_id=eid,
                    checklist_json=json.dumps(checklist),
                    verified_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ))
        db.session.commit()
        return jsonify({"status": "success", "message": f"{len(payload)} pallet(s) saved."})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/delete_pallet_verification/<sr_no>/<path:element_id>', methods=['DELETE'])
def delete_pallet_verification(sr_no, element_id):
    try:
        pv = PalletVerification.query.filter_by(sr_no=sr_no, element_id=element_id).first()
        if pv:
            db.session.delete(pv)
            db.session.commit()
        return jsonify({"status": "success", "message": f"'{element_id}' removed."})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


# ══════════════════════════════════════════════════════
#  ROUTES — Step 4: Photos
# ══════════════════════════════════════════════════════

@app.route('/photos/<sr_no>')
def photos_page(sr_no):
    entry_data = ShippingEntry.query.filter_by(sr_no=sr_no).first_or_404()
    photos     = ShipmentPhoto.query.filter_by(sr_no=sr_no).order_by(ShipmentPhoto.id).all()
    return render_template('photos.html', data=entry_data, photos=photos)

@app.route('/upload_photo/<sr_no>', methods=['POST'])
def upload_photo(sr_no):
    entry_data = ShippingEntry.query.filter_by(sr_no=sr_no).first_or_404()

    if 'photo' not in request.files:
        return jsonify({"status": "error", "message": "No file part"}), 400

    file = request.files['photo']
    if file.filename == '':
        return jsonify({"status": "error", "message": "No file selected"}), 400

    try:
        today       = datetime.now().strftime('%d-%m-%Y')
        raw_name    = f"{today}_{entry_data.sr_no}_{entry_data.customer_name or 'UNKNOWN'}"
        folder_name = re.sub(r'[^\w\s\-]', '', raw_name).strip()

        folder_path = os.path.join(basedir, 'static', 'app data', folder_name, 'Loading Photos')
        os.makedirs(folder_path, exist_ok=True)

        ts        = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:19]
        ext       = os.path.splitext(secure_filename(file.filename))[1] or '.jpg'
        filename  = f"{sr_no}_{ts}{ext}"
        file.save(os.path.join(folder_path, filename))

        rel_path = os.path.join('app data', folder_name, 'Loading Photos', filename).replace('\\', '/')

        photo = ShipmentPhoto(
            sr_no=sr_no, file_path=rel_path,
            category='Loading Photos',
            uploaded_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        db.session.add(photo)
        db.session.commit()

        return jsonify({"status": "success", "photo_id": photo.id,
                        "url": f"/static/{rel_path}", "path": rel_path})

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/delete_photo/<int:photo_id>', methods=['DELETE'])
def delete_photo(photo_id):
    photo = ShipmentPhoto.query.get_or_404(photo_id)
    try:
        full_path = os.path.join(basedir, 'static', photo.file_path)
        if os.path.exists(full_path):
            os.remove(full_path)
        db.session.delete(photo)
        db.session.commit()
        return jsonify({"status": "success", "message": "Photo deleted."})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


# ══════════════════════════════════════════════════════
#  STARTUP
# ══════════════════════════════════════════════════════

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)